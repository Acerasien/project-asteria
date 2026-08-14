import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Templat Impor Tamu"

    # Define headers
    headers = [
        "Nama Lengkap",
        "Jenis Kelamin",
        "Telepon",
        "Email",
        "Nomor Identitas",
        "Catatan"
    ]
    
    ws.append(headers)
    
    # Define sample data
    samples = [
        ["Budi Santoso", "Laki-laki", "+628123456789", "budi@example.com", "1234567890123456", "Tamu VIP"],
        ["Siti Aminah", "Perempuan", "+628987654321", "siti@example.com", "3210987654321098", "Alergi kacang"]
    ]
    
    for row in samples:
        ws.append(row)
        
    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark slate matching brand
    alignment = Alignment(horizontal="left", vertical="center")
    
    # Format header row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        
    # Format data rows
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    for row_idx in range(2, len(samples) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=11)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
    # Set column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Adjust row heights
    ws.row_dimensions[1].height = 25
    for r in range(2, len(samples) + 2):
        ws.row_dimensions[r].height = 20
        
    # Ensure directory exists
    os.makedirs("public", exist_ok=True)
    wb.save("public/templat_impor_tamu.xlsx")
    print("Excel template generated successfully at public/templat_impor_tamu.xlsx")

if __name__ == "__main__":
    generate_template()
